"""
文档读取工具

从项目策划文档目录读取文件内容。
文档根目录由 .env 中的 PROJECT_DOC_PATH 控制。
"""

import os
from pathlib import Path

from dotenv import load_dotenv

load_dotenv()


def get_doc_base_path() -> Path:
    return Path(os.getenv("PROJECT_DOC_PATH", r""))


def list_directory(subdir: str = "") -> str:
    """列出文档目录下的文件和子目录。"""
    base = get_doc_base_path()
    target = base / subdir if subdir else base
    if not target.exists():
        return f"目录不存在: {subdir or '（根目录）'}"

    items = sorted(target.iterdir(), key=lambda p: (p.is_file(), p.name))
    lines = []
    for item in items:
        prefix = "📄" if item.is_file() else "📁"
        lines.append(f"{prefix} {item.name}{'/' if item.is_dir() else ''}")
    return "\n".join(lines) if lines else "（空目录）"


def read_file(relative_path: str) -> str:
    """读取文档文件内容。支持 .xlsx、.docx、.md、.txt。"""
    base = get_doc_base_path()
    full_path = base / relative_path

    if not full_path.exists():
        return f"错误：文件不存在 - {relative_path}"

    suffix = full_path.suffix.lower()

    if suffix in (".xlsx", ".xls"):
        try:
            import openpyxl
            wb = openpyxl.load_workbook(full_path, read_only=True, data_only=True)
            result = []
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                result.append(f"### Sheet: {sheet_name}")
                for row in ws.iter_rows(values_only=True):
                    cells = [str(c) if c is not None else "" for c in row]
                    if any(c.strip() for c in cells):
                        result.append(" | ".join(cells))
            wb.close()
            return "\n".join(result)
        except ImportError:
            return "需要安装 openpyxl：pip install openpyxl"

    elif suffix in (".md", ".txt"):
        return full_path.read_text(encoding="utf-8")

    else:
        return f"不支持的文件类型：{suffix}（支持 .xlsx/.xls/.docx/.md/.txt）"
