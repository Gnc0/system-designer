"""
xlsx_to_md.py
只读取 xlsx 中的"功能需求" Sheet，转换为 md 文档（制表符缩进内容，无 Markdown 标题/列表语法）。
- 不加载图片（read_only=True），不浪费 token。
- 只处理 TARGET_SHEET，其他 sheet 一律跳过，避免竞品/UI/配置内容混淆上下文。
- 通过列号推断层级：col1=一级标题(0\t), col2=二级标题(1\t), col3=三级(2\t), col4+=内容项(n-1个\t)
- 同一行多列内容用 \t 合并（方便复制到 Excel）。

用法：
    python src/xlsx_to_md.py <xlsx_path> [output_dir]

示例：
    python src/xlsx_to_md.py "D:/Docs/策划/系统文档/C-成就/C-成就.xlsx" docs/reference
"""

import sys
import os
import openpyxl

# 优先读取此 Sheet；若不存在，退而使用模糊匹配（见 find_target_sheet）
TARGET_SHEET = "功能需求"


def find_target_sheet(sheetnames: list, xlsx_path: str) -> str | None:
    """
    按优先级查找目标 Sheet：
    1. 精确匹配 TARGET_SHEET（"功能需求"）
    2. 从文件名提取系统名（去掉 'X-' 前缀），对所有页签做包含匹配（模糊）
    返回匹配到的 Sheet 名；未找到则返回 None。
    """
    # 优先精确匹配
    if TARGET_SHEET in sheetnames:
        return TARGET_SHEET

    # 从文件名提取系统名，例如 "C-材料本优化.xlsx" → "材料本优化"
    base = os.path.splitext(os.path.basename(xlsx_path))[0]
    # 去掉形如 "X-" 的前缀（单字母 + 连字符）
    import re
    system_name = re.sub(r"^[A-Za-z]-", "", base).strip()

    if not system_name:
        return None

    # 模糊匹配：页签名包含系统名，或系统名包含页签名
    for name in sheetnames:
        if system_name in name or name in system_name:
            return name

    return None


def col_to_txt(min_col: int, text: str) -> str:
    """根据列号返回对应 txt 行（制表符缩进，无 Markdown 语法）。
    层级规则：level n → (n-1) 个 \t，一级标题前后加空行。
    """
    if min_col <= 1:
        return f"\n{text}\n"
    tabs = "\t" * (min_col - 1)
    return f"{tabs}{text}"


def sheet_to_txt(sheet_name: str, rows: list) -> str:
    """将一个 sheet 的行数据转为 txt 文本块（制表符缩进，无 Markdown 语法）。"""
    if not rows:
        return ""

    lines = [f"{sheet_name}\n"]

    for row in rows:
        if not row:
            continue

        # 按列号排序，获取最小列（决定层级）
        row_sorted = sorted(row, key=lambda c: c["col"])
        min_col = row_sorted[0]["col"]

        # 合并同行所有文本（多列用 \t 分隔，方便粘贴到 Excel）
        texts = [c["val"].strip() for c in row_sorted if c["val"].strip()]
        if not texts:
            continue
        combined = "\t".join(texts)

        lines.append(col_to_txt(min_col, combined))

    return "\n".join(lines)


def xlsx_to_md(xlsx_path: str, output_dir: str) -> str:
    """
    只读取 TARGET_SHEET（功能需求），写入 txt 文件（制表符缩进格式）。
    返回输出文件路径。
    """
    os.makedirs(output_dir, exist_ok=True)

    base_name = os.path.splitext(os.path.basename(xlsx_path))[0]
    output_path = os.path.join(output_dir, f"{base_name}.md")

    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)

    matched_sheet = find_target_sheet(wb.sheetnames, xlsx_path)
    if matched_sheet is None:
        all_sheets = wb.sheetnames
        wb.close()
        print(f"[ERROR] 未找到 '{TARGET_SHEET}' 或匹配的页签（{xlsx_path}）")
        print(f"        可用页签：{all_sheets}")
        sys.exit(1)

    if matched_sheet != TARGET_SHEET:
        print(f"[INFO] 未找到 '{TARGET_SHEET}'，模糊匹配使用页签：'{matched_sheet}'")

    ws = wb[matched_sheet]
    rows = []

    for row in ws.iter_rows():
        row_data = []
        for cell in row:
            v = cell.value
            if v is not None:
                text = str(v).strip()
                if text:
                    row_data.append({"row": cell.row, "col": cell.column, "val": text})
        if row_data:
            rows.append(row_data)

    wb.close()

    if not rows:
        print(f"[WARNING] '{TARGET_SHEET}' sheet is empty in {xlsx_path}")
        return output_path

    section_txt = sheet_to_txt(matched_sheet, rows)
    header = f"{base_name}\n\n由 xlsx_to_md.py 自动生成，源文件：{os.path.basename(xlsx_path)}，Sheet：{matched_sheet}\n\n"
    full_content = header + section_txt

    with open(output_path, "w", encoding="utf-8") as f:
        f.write(full_content)

    print(f"[OK] 已生成: {output_path}")
    return output_path


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("用法: python src/xlsx_to_md.py <xlsx_path> [output_dir]")
        sys.exit(1)

    xlsx_path = sys.argv[1]
    output_dir = sys.argv[2] if len(sys.argv) > 2 else "docs/reference"

    xlsx_to_md(xlsx_path, output_dir)
